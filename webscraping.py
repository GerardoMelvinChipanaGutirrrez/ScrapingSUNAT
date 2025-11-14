import time
import pandas as pd
import re
import os
import glob
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, TimeoutException

# ---------------- CONFIGURACIÓN ----------------
escritorio = r"F:\Programas sin finalizar\ScrapingSUNAT"

# Buscar archivos que empiecen con "empresas_" y terminen en ".xlsx"
patron = os.path.join(escritorio, "empresas_*.xlsx")
coincidencias = glob.glob(patron)

if coincidencias:
    ruta_entrada = coincidencias[0]
    print(f"✔ Archivo encontrado dinámicamente: {ruta_entrada}")
else:
    # fallback a empresas.xlsx si no hay empresas_*.xlsx
    ruta_entrada = os.path.join(escritorio, "empresas.xlsx")
    print(f" No se encontró 'empresas_*.xlsx'. Usando archivo por defecto: {ruta_entrada}")

ruta_salida = os.path.join(escritorio, "resultados_sunat_detallado.xlsx")

opts = Options()
opts.add_argument("--start-maximized")
opts.add_argument("--disable-blink-features=AutomationControlled")
opts.add_argument("--disable-infobars")
opts.add_argument("--disable-gpu")
opts.add_argument("--no-sandbox")
opts.add_argument("--disable-dev-shm-usage")
opts.add_argument("--ignore-certificate-errors")
opts.add_experimental_option("excludeSwitches", ["enable-automation"])
opts.add_experimental_option('useAutomationExtension', False)
opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36")

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=opts)
wait = WebDriverWait(driver, 10)


# ---------------- FUNCIÓN PRINCIPAL (DEBUG PASO A PASO) ----------------
def buscar_en_sunat(nombre=None, ruc=None, intentos=2):
    """
    Versión mejorada: detecta automáticamente si la búsqueda es por RUC o por nombre.
    Si es por RUC, salta la parte de 'div.list-group a' y va directo al detalle.
    """
    for intento in range(1, intentos + 1):
        try:
            print(f"\n=== Intento {intento} - Buscando: {nombre or ruc} ===")
            driver.get("https://e-consultaruc.sunat.gob.pe/cl-ti-itmrconsruc/FrameCriterioBusquedaWeb.jsp")
            time.sleep(0.8)

            # --- Selección de búsqueda por nombre o RUC ---
            if nombre:
                wait.until(EC.element_to_be_clickable((By.ID, "btnPorRazonSocial"))).click()
                input_nombre = wait.until(EC.presence_of_element_located((By.ID, "txtNombreRazonSocial")))
                input_nombre.clear()
                input_nombre.send_keys(nombre)
                print(f"  -> Nombre ingresado: {nombre}")
            elif ruc:
                wait.until(EC.element_to_be_clickable((By.ID, "btnPorRuc"))).click()
                input_ruc = wait.until(EC.presence_of_element_located((By.ID, "txtRuc")))
                input_ruc.clear()
                input_ruc.send_keys(ruc)
                print(f"  -> RUC ingresado: {ruc}")
            else:
                print("❌ Ningún parámetro (nombre/ruc) proporcionado.")
                return None

            # --- Click en Aceptar ---
            wait.until(EC.element_to_be_clickable((By.ID, "btnAceptar"))).click()
            time.sleep(1.5)

            # --- Si la búsqueda fue por NOMBRE ---
            if nombre and not ruc:
                try:
                    resultados = wait.until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.list-group a"))
                    )
                    print(f"🔍 Se encontraron {len(resultados)} posibles resultados.")
                except TimeoutException:
                    print("⚠️ Timeout: no se encontraron elementos 'div.list-group a' (probablemente sin resultados).")
                    # Buscar en Google (universidadperu.com)
                    ruc_google = buscar_ruc_en_universidad_peru(nombre)
                    if ruc_google:
                        print(f"🔁 Reintentando búsqueda en SUNAT por RUC: {ruc_google}")
                        resultado_ruc = buscar_en_sunat(ruc=ruc_google)
                        return resultado_ruc
                    continue

                # Buscar coincidencia exacta por nombre
                elegido = None
                for idx, a in enumerate(resultados, start=1):
                    try:
                        h4s = a.find_elements(By.TAG_NAME, "h4")
                        texto_h4 = h4s[1].text.strip().upper() if len(h4s) >= 2 else ""
                        print(f"    Enlace #{idx}: segundo h4 -> '{texto_h4[:80]}'")
                        if nombre.strip().upper() in texto_h4:
                            elegido = a
                            print(f"    ✅ Coincidencia encontrada en enlace #{idx}.")
                            break
                    except Exception as e:
                        print(f"    ⚠️ Error leyendo h4 del enlace #{idx}: {e}")

                # Si no encontró coincidencia, probar con Google
                if not elegido:
                    print(f"⚠️ No se encontró coincidencia exacta para {nombre}. Buscando RUC en universidadperu.com...")
                    ruc_google = buscar_ruc_en_universidad_peru(nombre)
                    if ruc_google:
                        print(f"🔁 Reintentando búsqueda en SUNAT por RUC: {ruc_google}")
                        resultado_ruc = buscar_en_sunat(ruc=ruc_google)
                        return resultado_ruc
                    else:
                        continue

                # Clic en el enlace elegido
                try:
                    elegido.click()
                except Exception as e_click:
                    print(f"  ⚠️ Clic normal falló: {e_click} -> intento con JavaScript.")
                    try:
                        driver.execute_script("arguments[0].click();", elegido)
                        print("  -> Clic via JS ejecutado.")
                    except Exception as e_js:
                        print(f"  ❌ Clic via JS también falló: {e_js}")
                        continue

                # Esperar detalle
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.list-group-item")))

            # --- Si la búsqueda fue por RUC ---
            else:
                print("🧾 Búsqueda por RUC detectada. Saltando lista de resultados.")
                wait.until(EC.presence_of_element_located((By.XPATH, "//h4[contains(text(),'Número de RUC')]")))

            time.sleep(0.6)

            # --- Extraer RUC y nombre de la empresa ---
            try:
                # Esperamos que el primer div esté presente y visible
                primer_div = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.list-group-item")))
                time.sleep(0.5)  # Pequeña espera para asegurar que el contenido esté cargado
                
                # Obtener el texto completo del div que contiene RUC y nombre
                texto_completo = primer_div.text
                print(f"  -> Texto completo del div: '{texto_completo[:120]}'")
                
                # Extraer RUC y nombre usando regex
                match_ruc_nombre = re.search(r"(\d{11})\s*-\s*(.+)", texto_completo)
                if match_ruc_nombre:
                    ruc_numero = match_ruc_nombre.group(1)
                    nombre_empresa = match_ruc_nombre.group(2).strip()
                    print(f"  ✓ RUC extraído: {ruc_numero}")
                    print(f"  ✓ Nombre extraído: {nombre_empresa}")
                else:
                    print("  ⚠️ No se pudo extraer RUC/nombre del texto")
                    ruc_numero = ""
                    nombre_empresa = texto_completo
            except Exception as e:
                print(f"  ⚠️ Error al extraer RUC/nombre: {e}")
                ruc_numero = ""
                nombre_empresa = ""

            # --- Verificar estado activo/habido ---
            divs_verdes = driver.find_elements(By.CSS_SELECTOR, "div.list-group-item.list-group-item-success")
            if len(divs_verdes) < 2:
                print(f"⚠️ Empresa {nombre or ruc} no parece activa/habida (se considera baja).")
                return "baja", "baja", []

            # --- Obtener representantes legales ---
            representantes = []
            try:
                try:
                    btn_rep = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Representante')]")))
                    btn_rep.click()
                    time.sleep(0.8)
                except Exception:
                    print("  -> No se encontró o no se pudo clicar el botón 'Representante' (se intentará leer la tabla).")

                filas_rep = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
                for i, fila in enumerate(filas_rep, start=1):
                    try:
                        celdas = fila.find_elements(By.TAG_NAME, "td")
                        if len(celdas) >= 4:
                            dni = celdas[0].text.strip()
                            nombre_rep = celdas[1].text.strip()
                            cargo = celdas[2].text.strip()
                            fecha = celdas[3].text.strip()
                            representantes.append((dni, nombre_rep, cargo, fecha))
                            print(f"    -> Rep #{i}: {dni}, {nombre_rep}, {cargo}, {fecha}")
                        else:
                            print(f"    -> Fila #{i} ignorada (menos de 4 celdas).")
                    except Exception as e_fila:
                        print(f"    ⚠️ Error procesando fila #{i}: {e_fila}")
            except Exception as e_rep:
                print(f"  ⚠️ Error al extraer representantes: {e_rep}")
                representantes = []

            print(f"=== FIN intento {intento} - éxito ===")
            return ruc_numero, nombre_empresa, representantes

        except Exception as e:
            print(f"❌ Error inesperado en intento {intento} para {nombre or ruc}: {e}")
            time.sleep(1.5)
            continue

    print(f"❌ Falló la búsqueda definitiva para {nombre or ruc} después de {intentos} intentos.")
    return None


def buscar_ruc_en_universidad_peru(nombre):
    """
    Busca en Google con la consulta "<nombre> ruc universidad peru"
    y extrae el RUC directamente desde la página de universidadperu.com.
    """
    try:
        consulta = f"{nombre} ruc universidad peru"
        print(f"🌐 Buscando en Google: {consulta}")
        driver.get("https://www.google.com")
        time.sleep(1)

        # Aceptar cookies si aparece el botón (solo la primera vez)
        try:
            btn_aceptar = driver.find_element(By.XPATH, "//button//*[text()='Aceptar todo']/..")
            btn_aceptar.click()
            time.sleep(1)
        except:
            pass  # Puede que no aparezca

        # Buscar en Google
        caja = wait.until(EC.presence_of_element_located((By.NAME, "q")))
        caja.clear()
        caja.send_keys(consulta)
        caja.submit()
        time.sleep(2.5)

        # Buscar el primer resultado de universidadperu.com
        enlaces = driver.find_elements(By.CSS_SELECTOR, "a h3")
        url_universidad_peru = None

        for enlace in enlaces:
            try:
                a_tag = enlace.find_element(By.XPATH, "./ancestor::a")
                href = a_tag.get_attribute("href")
                if "universidadperu.com" in href:
                    url_universidad_peru = href
                    print(f"  🔗 Enlace encontrado: {href}")
                    break
            except Exception:
                continue

        if not url_universidad_peru:
            print("⚠️ No se encontró ningún resultado de universidadperu.com en Google.")
            return None

        # Entrar a la página de universidadperu.com
        driver.get(url_universidad_peru)
        time.sleep(2)

        # Extraer el texto completo
        texto = driver.find_element(By.TAG_NAME, "body").text

        # Buscar patrón que contenga RUC
        # Ejemplo de formato en la web: RUC: 20603684291
        match = re.search(r"RUC[:\s]+(\d{11})", texto)
        if match:
            ruc_encontrado = match.group(1)
            print(f"  ✅ RUC encontrado en universidadperu.com: {ruc_encontrado}")
            return ruc_encontrado
        else:
            # Intento alternativo: buscar cualquier número de 11 dígitos
            match_alt = re.search(r"\b(\d{11})\b", texto)
            if match_alt:
                ruc_encontrado = match_alt.group(1)
                print(f"  ✅ RUC (por coincidencia numérica): {ruc_encontrado}")
                return ruc_encontrado
            else:
                print("⚠️ No se encontró ningún número de RUC en la página.")
                return None

    except Exception as e:
        print(f"❌ Error durante la búsqueda del RUC en universidadperu.com: {e}")
        return None



# ---------------- LECTURA DEL EXCEL ----------------
try:
    df = pd.read_excel(ruta_entrada)
except Exception as e:
    print(f"⚠️ Error leyendo el archivo Excel: {e}")
    driver.quit()
    exit()

# ---------------- CREAR ARCHIVO DE SALIDA VACÍO ----------------
from openpyxl import Workbook, load_workbook

if not os.path.exists(ruta_salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append([
        "razon_social", "ruc", "nombre_encontrado",
        "dni_representante", "nombre_representante",
        "cargo", "fecha_designacion"
    ])
    wb.save(ruta_salida)
    print(f"📘 Archivo creado: {ruta_salida}")
else:
    print(f"📘 Archivo existente detectado: {ruta_salida}")

# ---------------- PROCESAR CADA EMPRESA ----------------
for razon in df["razon_social"]:
    print(f"\n🔍 Buscando: {razon}")
    data = buscar_en_sunat(nombre=razon)

    filas_a_agregar = []

    if data:
        ruc, nombre_empresa, reps = data
        if ruc == "baja":
            filas_a_agregar.append([
                "baja", razon, "baja", "", "", "", ""
            ])
        elif reps:
            for dni, nombre_rep, cargo, fecha in reps:
                filas_a_agregar.append([
                    ruc, razon, nombre_rep, cargo, 
                ])
        else:
            filas_a_agregar.append([
                razon, ruc, nombre_empresa, "", "", "", ""
            ])
    else:
        filas_a_agregar.append([
            razon, "", "No encontrado", "", "", "", ""
        ])

    # ---------------- GUARDAR RESULTADOS EN EL EXCEL ----------------
    try:
        wb = load_workbook(ruta_salida)
        ws = wb.active
        for fila in filas_a_agregar:
            ws.append(fila)
        wb.save(ruta_salida)
        print(f"💾 Datos guardados para: {razon}")
    except Exception as e:
        print(f"⚠️ Error guardando datos para {razon}: {e}")

print(f"\n✅ Proceso completado. Los resultados se fueron guardando en tiempo real en:\n{ruta_salida}")

driver.quit()