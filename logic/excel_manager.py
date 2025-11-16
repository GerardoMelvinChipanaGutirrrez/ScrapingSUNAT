import pandas as pd
from openpyxl import Workbook, load_workbook

def leer_excel(ruta):
    return pd.read_excel(ruta)

def crear_excel_salida(ruta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append([
        "razon_social", "ruc", "nombre_empresa",
        "dni_representante", "nombre_representante",
        "cargo", "fecha_designacion"
    ])
    wb.save(ruta)

def agregar_fila(ruta, fila):
    wb = load_workbook(ruta)
    ws = wb.active
    ws.append(fila)
    wb.save(ruta)
